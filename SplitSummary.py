import calendar
import datetime
import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font,Border,Side,PatternFill
from openpyxl.utils import get_column_letter


class DashJasper:
    def __init__(self,report_path,client,report_name,report_date):

        self.location = set()
        self.reports = report_name
        self.season = set()
        # self.month = set()
        self.report_date = report_date
        self.client = client
        self.last_report = report_name[-1]
        self.report = report_path
        self.loc_data = self.season_data = []
        self.summaryheader = ["GL Code","Name","Amount","No."]
        self.headername = "Unearned Revenue Summary as of"
        self.gl_num = self.sdate = self.edate = None
        self.summary = {}
        self.summary_final = []
        self.count_line = 0


    def load_summary(self):
        loc_num = season_num = mon_num = None
        summary_wb = None
        try:
            summary_wb = load_workbook(self.report,data_only=True,read_only=True)
        except Exception as e:
            print(str(e))
        # self.reports = ["Camp"]
        for report_num,report in enumerate(self.reports):

            summary_ws = summary_wb[report].values
            self.summary_data = [list(row) for row in summary_ws if row]
            if "Event & Rental" not in report:
                for row_num,row in enumerate(self.summary_data):
                    if row_num == 0:
                        for header_num,header_row in enumerate(row):
                            if header_row == "Location":
                                self.loc_num = header_num
                            elif header_row == "Season Name":
                                self.season_num = header_num
                            # elif header_row == "Month":
                            #     self.mon_num = header_num
                            elif header_row == "GL Code":
                                self.gl_num = header_num
                            elif header_row == "Start Date":
                                self.sdate = header_num
                            elif header_row == "End Date":
                                self.edate = header_num
                            elif header_row == "Unearned Revenue Closing":
                                self.amount = header_num
                        continue
                    self.location.add(row[self.loc_num])
                    self.season.add(row[self.season_num])
                    # self.month.add(row[self.mon_num])
            else:
                for row_num, row in enumerate(self.summary_data):
                    if row_num == 0:
                        for header_num, header_row in enumerate(row):
                            if header_row == "Location":
                                loc_num = header_num
                            elif header_row == "GL Code":
                                self.egl_num = header_num
                            elif header_row == "Event Date":
                                self.esdate = header_num
                            elif header_row == "Invoice Date":
                                self.eedate = header_num
                            elif header_row == "GLCode Description":
                                self.description = header_num
                            elif header_row == "Amount":
                                self.eamount = header_num

                        continue
                    self.location.add(row[loc_num])


            summary = self.processed_summary(report)
            if not summary:
                return False
        return True

    def processed_summary(self,report_name):
        file_path = None
        class_count = camp = event = 0

        path = os.path.join(os.getcwd(),"Downloads",self.client,str(self.report_date))
        if not os.path.exists(path):
            os.makedirs(path)
        sheetname = None
        for loc_num,loc in enumerate(self.location):

            total = 0.00
            count = 1
            loc_wb = Workbook()
            loc_ws = loc_wb.active
            loc_ws.title = "Summary"
            temp_ws = loc_ws
            self.header_row = self.loc_data = []
            loc_path = os.path.join(path,loc)
            if not os.path.exists(loc_path):
                os.makedirs(loc_path)
            file_path = os.path.join(loc_path,f'{loc}.xlsx')
            for row_num,row in enumerate(self.summary_data):
                if row_num == 0:
                    self.header_row = list(row)
                    continue
                if row[0] == loc:
                    self.loc_data.append(row)

            if "Event & Rental" not in report_name:
                for season_num,season in enumerate(self.season):
                    for row_num,row in enumerate(self.loc_data):
                        if season == row[3]:
                            self.season_data.append(row)

                    for data_num, data in enumerate(self.season_data):
                        if data:
                            data[self.amount] = float(
                                str(data[self.amount]).replace("$", "").replace(",", "").replace(" ", "").strip())

                            if data_num == 0:
                                if os.path.exists(file_path):
                                   loc_wb = load_workbook(file_path)
                                sheetname = str(report_name[:2]).upper()
                                sheetname = f'{sheetname}{count}'
                                loc_ws = loc_wb.create_sheet(sheetname)
                                loc_ws.sheet_view.showGridLines = False
                                count += 1
                                loc_ws.append(self.header_row)

                                loc_ws.append(data)
                            else:
                                loc_ws.append(data)
                            loc_wb.save(file_path)
                            amount = data[self.amount]
                            if self.summary:
                                if (data[self.gl_num], data[self.season_num], sheetname) in self.summary.keys():
                                    self.summary[
                                        data[self.gl_num], data[self.season_num], sheetname] += amount
                                else:
                                    self.summary[
                                        data[self.gl_num], data[self.season_num], sheetname] = amount
                            else:
                                self.summary[
                                    data[self.gl_num], data[self.season_num], sheetname] = amount
                            for i in range(self.sdate+1, self.edate+2):
                                loc_ws.cell(row=data_num+1, column=i).number_format = f'mm-dd-yyyy'

                            if report_name == "Class":
                                for x in range(10, 15):
                                    loc_ws.cell(row=data_num+2, column=x).number_format = f'$#,##0.00_);[Red]($#,##0.00)'
                            elif report_name == "Camp":
                                for j in range(11, 20):
                                    loc_ws.cell(row=data_num+2, column=j).number_format = f'$#,##0.00_);[Red]($#,##0.00)'
                            loc_wb.save(file_path)


                    self.season_data.clear()
            else:
                for data_num,data in enumerate(self.loc_data):
                    if data:

                        if data_num == 0:
                            if os.path.exists(file_path):
                                loc_wb = load_workbook(file_path)
                            sheetname = f'{str(report_name[:2]).upper()}{count}'
                            loc_ws = loc_wb.create_sheet(sheetname)
                            loc_ws.sheet_view.showGridLines = False
                            count += 1
                            loc_ws.append(self.header_row)

                            loc_ws.append(data)
                        else:
                            loc_ws.append(data)
                        data[self.eamount] = float(
                            str(data[self.eamount]).replace("$", "").replace(",", "").replace(" ", "").strip())
                        if self.summary:
                            if (data[self.egl_num], data[self.description], sheetname) in self.summary.keys():
                                self.summary[data[self.egl_num], data[self.description], sheetname] += \
                                    data[self.eamount]
                            else:
                                self.summary[data[self.egl_num], data[self.description], sheetname] = \
                                    data[self.eamount]
                        else:
                            self.summary[data[self.egl_num], data[self.description], sheetname] = data[
                                self.eamount]
                        for k in range(self.esdate+1,self.eedate+2):
                            if k == 7:
                                continue
                            loc_ws.cell(row=data_num+2, column=k).number_format = f'mm-dd-yyyy'
                        for p in range(9,11):
                            loc_ws.cell(row=data_num + 2, column=p).number_format = f'$#,##0.00_);[Red]($#,##0.00)'
                loc_wb.save(file_path)
            if self.summary_final:
                self.summary_final.clear()
                self.summary_final.append([])
                self.summary_final.append([])
            self.summary_final.append([None,None,f'{report_name} - Unearned Revenue'])
            self.summary_final.append([None, "GL Code", "Name", "Amount", "No"])
            if not self.summary:
                continue
            for key, value in self.summary.items():
                if self.summary:
                    if value:
                        glcode = key[0]
                        name = key[1]
                        sname = key[2]
                        total += value
                        self.summary_final.append([None,glcode,name,value,sname])
            self.summary.clear()
            self.summary_final.append([])
            if total:
                self.summary_final.append([None,None,"Total",total,None])

            loc_wb = load_workbook(file_path)
            loc_ws = loc_wb["Summary"]
            loc_ws.sheet_view.showGridLines = False

            for row in self.summary_final:
                loc_ws.append(row)
            loc_wb.save(file_path)

            # loc_ws.append([None,None,"Grand Total",grand_total,None])
            # loc_wb.save(file_path)
            grand_total = 0
            xltotal = 0
            last_row = loc_ws.max_row
            last_col = loc_ws.max_column
            for row in range(1,last_row+1):
                for col_num in range(2,last_col+1):

                    # if row >= 4:
                    xlvalue = loc_ws.cell(row=row, column=3).value or 0
                    if xlvalue:
                        if 'Unearned Revenue' not in xlvalue:
                            if xlvalue == "Total":
                                loc_ws.cell(row=row - 1, column=col_num).border = \
                                    Border(top=Side(border_style='thin', color='FF000000'),
                                           right=Side(border_style='thin', color='FF000000'),
                                           left=Side(border_style='thin', color='FF000000'),
                                           bottom=Side(border_style='thin', color='FF000000'))
                                loc_ws.cell(row=row, column=col_num).border = \
                                    Border(top=Side(border_style='thin', color='FF000000'),
                                           right=Side(border_style='thin', color='FF000000'),
                                           left=Side(border_style='thin', color='FF000000'),
                                           bottom=Side(border_style='thin', color='FF000000'))
                            else:
                                loc_ws.cell(row=row, column=col_num).border = \
                                    Border(top=Side(border_style='thin', color='FF000000'),
                                           right=Side(border_style='thin', color='FF000000'),
                                           left=Side(border_style='thin', color='FF000000'),
                                           bottom=Side(border_style='thin', color='FF000000'))



                        if xlvalue == 'Total' and (col_num == 3 or col_num == 4):
                            loc_ws.cell(row=row, column=col_num).font = Font(bold=True)
                            loc_ws.cell(row=row, column=col_num + 1).font = Font(bold=True)
                            loc_ws.cell(row=row, column=col_num).border = \
                                Border(top=Side(border_style='thin', color='FF000000'),
                                       right=Side(border_style='thin', color='FF000000'),
                                       left=Side(border_style='thin', color='FF000000'),
                                       bottom=Side(border_style='double', color='FF000000'))

                    xlvalue = loc_ws.cell(row=row,column=col_num).value
                    if row == 2 and col_num == 2:
                        if loc_ws.cell(row=2, column=2).value == None:
                            loc_ws.insert_rows(2)
                            loc_ws.merge_cells(start_row=2,end_row=2,start_column=2,end_column=5)
                            loc_ws.cell(row=2, column=2).alignment = Alignment(horizontal='center', vertical='center')
                            loc_ws.cell(row=2, column=2).font = Font(name="Calibri", bold=True, size=16,
                                                                           color="FFFFFF")
                            loc_ws.cell(row=2, column=2).fill = PatternFill(start_color='FF00B0F0',
                                                                                               end_color='FF00B0F0',
                                                                                               fill_type="solid")

                            loc_ws.cell(row=2, column=2).value = f'{self.headername} {str(self.report_date)}'



                    elif f'{report_name} - Unearned Revenue' == xlvalue:

                        loc_ws.cell(row=row, column=col_num).alignment = Alignment(horizontal='center',
                                                                                   vertical='center')
                        loc_ws.cell(row=row, column=col_num).font = Font(name="Calibri", bold=True, size=14,
                                                                         color="000000",underline='single')



                    elif "GL Code" == xlvalue or "Name" ==  xlvalue or "Amount" == xlvalue or "No" == xlvalue:
                        loc_ws.cell(row=row, column=col_num).alignment = Alignment(horizontal='center',
                                                                                   vertical='center')
                        loc_ws.cell(row=row, column=col_num).font = Font(name="Calibri",bold=True, size=11,
                                                                         color="000000")
                        loc_ws.cell(row=row, column=col_num).fill = PatternFill(start_color='FCE4D6',
                                                                            end_color='FCE4D6',
                                                                            fill_type="solid")
                        loc_ws.column_dimensions[get_column_letter(1)].width = 3
                        loc_ws.column_dimensions[get_column_letter(2)].width = 10
                        loc_ws.column_dimensions[get_column_letter(3)].width = 45
                        loc_ws.column_dimensions[get_column_letter(4)].width = 15
                        loc_ws.column_dimensions[get_column_letter(5)].width = 10



                    elif col_num == 4:
                        xlvalue = loc_ws.cell(row=row, column=4).value
                        if type(xlvalue) == float or type(xlvalue) == int:
                            loc_ws.cell(row=row, column=col_num).number_format = f'$#,##0.00_);[Red]($#,##0.00)'
                    elif row >= 3 and col_num == 5:
                        xlvalue = loc_ws.cell(row=row, column=5).value
                        if xlvalue != "No" and xlvalue:
                            link = f'#{xlvalue}!A1'
                            loc_ws.cell(row=row, column=5).hyperlink = link
                            loc_ws.cell(row=row,column=5).font = Font(underline='single', color='0563C1')
                            loc_ws.cell(row=row, column=5).alignment = Alignment(horizontal='center',
                                                                                       vertical='center')

                    grand_total = 0

                    if self.last_report in report_name and last_row == row:
                        for row_num in range(1,last_row+1):
                            xlvalue = loc_ws.cell(row=row_num,column=3).value
                            xltotal = loc_ws.cell(row=row_num, column=4).value
                            if xlvalue == 'Total' and xltotal:
                                grand_total += xltotal
                                loc_ws.cell(row=last_row + 3, column=3).value = "Total"
                                loc_ws.cell(row=last_row + 3, column=4).value = grand_total
                                loc_ws.cell(row=last_row + 3, column=3).border = \
                                    Border(top=Side(border_style='thin', color='FF000000'),
                                           bottom=Side(border_style='double', color='FF000000'))
                                loc_ws.cell(row=last_row + 3, column=4).border = \
                                    Border(top=Side(border_style='thin', color='FF000000'),
                                           bottom=Side(border_style='double', color='FF000000'))
                                loc_ws.cell(row=last_row + 3, column=4).number_format = f'$#,##0.00_);[Red]($#,##0.00)'
                                loc_ws.cell(row=last_row + 3, column=3).font = Font(name="Calibri", bold=True, size=11,
                                                                                    color="000000")
                                loc_ws.cell(row=last_row + 3, column=4).font = Font(name="Calibri", bold=True, size=11,
                                                                                    color="000000")
                loc_wb.save(file_path)
        return True


class RunDash:
    def __init__(self):
        self.gui_queue = None

    def run(self,report_path,client,report_name,report_date):
        dash = DashJasper(report_path,client,report_name,report_date)
        summary = dash.load_summary()
        if not summary:
            self.gui_queue.put({'status': 'Unable to process'}) if self.gui_queue else None
            return False
        self.gui_queue.put({'status':f'{client} reports processed successfully'}) if self.gui_queue else None
        return True




